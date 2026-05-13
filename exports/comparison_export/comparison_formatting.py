from exports.common.base_formatting import apply_base_formatting
from openpyxl.styles import PatternFill

def apply_comparison_formatting(worksheet, dataframe):
    """
    Applies formatting specific to the Comparison Export module.
    It builds on the base formatting but highlights CHANGES and COMMENTS columns.
    """
    # Base styling
    apply_base_formatting(worksheet, dataframe)
    
    # Highlight specific columns for comparison output
    changes_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    col_names_upper = [str(c).upper() for c in dataframe.columns]
    
    for row_num in range(2, len(dataframe) + 2):
        for col_num, col_name in enumerate(col_names_upper, start=1):
            if col_name in ("CHANGES", "COMMENTS"):
                worksheet.cell(row=row_num, column=col_num).fill = changes_fill
