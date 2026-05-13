import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def apply_base_formatting(worksheet, dataframe):
    """
    Applies baseline industry-standard Excel formatting to an openpyxl worksheet.
    Shared across different export pipelines.
    """
    # 1. Header Formatting
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    
    thin_border = Border(
        left=Side(border_style="thin", color="999999"),
        right=Side(border_style="thin", color="999999"),
        top=Side(border_style="thin", color="999999"),
        bottom=Side(border_style="thin", color="999999")
    )
    
    for col_num, col_name in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 2. Data Rows
    data_font = Font(name="Calibri", size=10)
    data_alignment_standard = Alignment(horizontal="left", vertical="top", wrap_text=False)
    data_alignment_numeric = Alignment(horizontal="center", vertical="top", wrap_text=False)
    data_alignment_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    numeric_columns = ["MIN", "TRGT", "MAX", "TRGT2", "QUANTITY", "TORQUE SNUG TARGET"]
    date_columns = ["DECISION DATE"]
    multiline_columns = ["COMMENTS", "CHANGES", "PART USAGE DESC", "PHYSCL DESC", "VSC NAME"]

    col_names_upper = [str(c).upper() for c in dataframe.columns]

    for row_num in range(2, len(dataframe) + 2):
        is_alt_row = (row_num % 2 == 0)

        for col_num, col_name in enumerate(col_names_upper, start=1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.border = thin_border

            if is_alt_row:
                cell.fill = alt_row_fill

            if col_name in multiline_columns:
                cell.alignment = data_alignment_wrap
            elif col_name in numeric_columns:
                cell.alignment = data_alignment_numeric
                cell.alignment = data_alignment_standard
                
            if col_name in numeric_columns:
                try:
                    if cell.value is not None:
                        float(cell.value)
                        cell.number_format = '0.00'
                except ValueError:
                    pass
            
            if col_name in date_columns:
                cell.number_format = 'YYYY-MM-DD'
                
    # 3. Auto Column Width
    for col_num, col_name in enumerate(col_names_upper, start=1):
        col_letter = get_column_letter(col_num)
        
        if col_name in multiline_columns:
            worksheet.column_dimensions[col_letter].width = 60
        else:
            max_len = len(str(dataframe.columns[col_num-1]))
            sample_data = dataframe.iloc[:50, col_num-1].astype(str)
            for val in sample_data:
                val_len = len(val.split('\n')[0]) if pd.notna(val) else 0
                if val_len > max_len:
                    max_len = val_len
            
            width = min(max_len + 2, 40)
            worksheet.column_dimensions[col_letter].width = width

    # 4. Freeze Header Row
    worksheet.freeze_panes = "A2"
    
    # 5. Enable Excel Filters
    last_col_letter = get_column_letter(len(dataframe.columns))
    filter_ref = f"A1:{last_col_letter}{len(dataframe) + 1}"
    worksheet.auto_filter.ref = filter_ref
    
    # 6. Row Height Adjustment
    worksheet.row_dimensions[1].height = 22
    for row_num in range(2, len(dataframe) + 2):
        max_newlines = 1
        for col_num, col_name in enumerate(col_names_upper, start=1):
            if col_name in multiline_columns:
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value and isinstance(cell_value, str):
                    newlines = cell_value.count('\n') + 1
                    if newlines > max_newlines:
                        max_newlines = newlines
        
        calculated_height = 18 + ((max_newlines - 1) * 14)
        worksheet.row_dimensions[row_num].height = min(calculated_height, 100)

    # 7. Zoom
    worksheet.sheet_view.zoomScale = 90
